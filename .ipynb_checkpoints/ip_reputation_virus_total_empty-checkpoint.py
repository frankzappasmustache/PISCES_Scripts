#!/usr/bin/env python3

import requests
import time
import openpyxl
from openpyxl.styles import Font, PatternFill
import ipaddress

# ── Configuration ───────────────────────────────────────────────────────────────
KIBANA_URL    = ''
KIBANA_USER   = ''
KIBANA_PASS   = ''

INDEX_PATTERN = '*'
VT_API_KEY    = ''
VT_URL        = 'https://www.virustotal.com/api/v3/ip_addresses/'

SEVERITIES    = [1, 2]
CLIENT_IDS    = ["", ""]
PER_RUN_LIMIT = 20
VT_PER_MINUTE = 4
DAILY_LIMIT   = 500

# ── Helpers ────────────────────────────────────────────────────────────────────

def is_public_ip(ip):
    try:
        return not ipaddress.ip_address(ip).is_private
    except Exception:
        return False

def check_ip_vt(ip):
    headers = {'x-apikey': VT_API_KEY}
    r = requests.get(f'{VT_URL}{ip}', headers=headers)
    if r.status_code == 200:
        data   = r.json()['data']['attributes']
        stats  = data['last_analysis_stats']
        # flag if any vendor is NOT harmless (i.e. not “Clean”)
        flagged = any(v['category'] != 'harmless'
                      for v in data['last_analysis_results'].values())
        return {
            'ip':         ip,
            'malicious':  stats['malicious'],
            'suspicious': stats['suspicious'],
            'reputation': data.get('reputation', 'N/A'),
            'flagged':    flagged
        }
    else:
        return {'ip': ip, 'error': r.json().get('error', {}).get('message', 'Unknown')}

# ── Kibana Login + Query ───────────────────────────────────────────────────────

session = requests.Session()
# Kibana’s security API expects a kbn-xsrf header even if you’re using basic auth
session.headers.update({'kbn-xsrf': 'true'})
session.auth = (KIBANA_USER, KIBANA_PASS)

def fetch_kibana_hits(size=10000):
    """
    Uses Kibana’s Console Proxy to run an ES _search, and handles both
    wrapped (with "responses") and unwrapped JSON formats.
    """
    es_query = {
      "query": {
        "bool": {
          "must": [
            {"terms": {"suricata.eve.alert.severity": SEVERITIES}},
            {"terms": {"clientID": CLIENT_IDS}},
            {"range": {"@timestamp": {"gte": "now-1h"}}},
            {"bool": {
               "should": [
                 {"bool": {"must_not": {"wildcard": {"source.address": "10.*"}}}},
                 {"bool": {"must_not": {"wildcard": {"source.address": "172.16.*"}}}},
                 {"bool": {"must_not": {"wildcard": {"source.address": "192.168.*"}}}},
                 {"bool": {"must_not": {"wildcard": {"destination.address": "10.*"}}}},
                 {"bool": {"must_not": {"wildcard": {"destination.address": "172.16.*"}}}},
                 {"bool": {"must_not": {"wildcard": {"destination.address": "192.168.*"}}}}
               ],
               "minimum_should_match": 1
            }}
          ]
        }
      },
      "size": size
    }

    proxy_path = f"/api/console/proxy?path={INDEX_PATTERN}/_search&method=POST"
    url        = KIBANA_URL.rstrip('/') + proxy_path
    resp       = session.post(url, json=es_query)
    resp.raise_for_status()

    data = resp.json()
    # Kibana console proxy sometimes wraps the ES response in "responses"
    if isinstance(data, dict) and "responses" in data:
        es_resp = data["responses"][0]
    else:
        es_resp = data

    return es_resp.get("hits", {}).get("hits", [])


# ── Extract, Limit, Excel ──────────────────────────────────────────────────────

def extract_ips(hits):
    ips = set()
    for hit in hits:
        src = hit['_source'].get('source',{}).get('ip')
        dst = hit['_source'].get('destination',{}).get('ip')
        for ip in (src, dst):
            if ip and is_public_ip(ip):
                ips.add(ip)
    return sorted(ips)[:PER_RUN_LIMIT]

def write_to_excel(results, filename='ip_reputation_results.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IP Reputation"
    ws.append(["IP", "Malicious", "Suspicious", "Reputation", "Error"])
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for r in results:
        ws.append([
            r['ip'],
            r.get('malicious','N/A'),
            r.get('suspicious','N/A'),
            r.get('reputation','N/A'),
            r.get('error','')
        ])
        if r.get('flagged'):
            for c in ws[ws.max_row]:
                c.fill = red_fill
                c.font = Font(bold=True)

    wb.save(filename)
    print(f"→ Saved to {filename}")

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("⏳ Fetching last 1 hr of alerts from Kibana…")
    hits = fetch_kibana_hits()
    print(f"→ {len(hits)} hits returned")

    ips = extract_ips(hits)
    print(f"→ {len(ips)} public IPs (capped at {PER_RUN_LIMIT})")

    results = []
    for idx, ip in enumerate(ips, start=1):
        if idx > DAILY_LIMIT:
            print("⚠️  Daily VT limit reached.")
            break
        r = check_ip_vt(ip)
        results.append(r)
        print(f"[{idx}/{len(ips)}] {ip} → malicious={r.get('malicious','?')}")
        if idx % VT_PER_MINUTE == 0:
            time.sleep(60)

    write_to_excel(results)

if __name__ == "__main__":
    main()
