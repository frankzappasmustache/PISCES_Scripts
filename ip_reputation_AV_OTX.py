#!/usr/bin/env python3
import os
import sys
import time
import requests
import openpyxl
import ipaddress
from openpyxl.styles import Font, PatternFill
from urllib.parse import urlparse

# ── Helpers ────────────────────────────────────────────────────────────────────
def ensure_url_scheme(u: str) -> str:
    p = urlparse(u)
    return u if p.scheme else 'https://' + u

def is_public_ip(ip: str) -> bool:
    try:
        return not ipaddress.ip_address(ip).is_private
    except Exception:
        return False

# ── Configuration ───────────────────────────────────────────────────────────────
KIBANA_URL        = ensure_url_scheme(os.getenv('KIBANA_URL','').strip())
KIBANA_USER       = os.getenv('KIBANA_USER','').strip()
KIBANA_PASS       = os.getenv('KIBANA_PASS','').strip()
INDEX_PATTERN     = '*'   # Elasticsearch index pattern

OTX_API_KEY       = os.getenv('OTX_API_KEY','').strip()
CLIENT_IDS        = [c for c in os.getenv('CLIENT_IDS','').split(',') if c]

# Throttling: distribute 10 000 daily queries over 24 runs (once/hour)
DAILY_QUERY_LIMIT = 10000
RUNS_PER_DAY      = 24
PER_RUN_LIMIT     = DAILY_QUERY_LIMIT // RUNS_PER_DAY  # ≈416 lookups/run

# ── Sanity checks ───────────────────────────────────────────────────────────────
missing = []
for var,name in [(KIBANA_URL,'KIBANA_URL'),
                 (KIBANA_USER,'KIBANA_USER'),
                 (KIBANA_PASS,'KIBANA_PASS'),
                 (OTX_API_KEY,'OTX_API_KEY'),
                 (CLIENT_IDS,'CLIENT_IDS')]:
    if not var:
        missing.append(name)
if missing:
    print(f"❌ ERROR: Missing environment variables: {', '.join(missing)}")
    sys.exit(1)

# ── OTX Query ──────────────────────────────────────────────────────────────────
def check_ip_otx(ip: str) -> dict:
    url     = f"https://otx.alienvault.com/api/v1/indicator/IPv4/{ip}/general"
    headers = {'X-OTX-API-KEY': OTX_API_KEY}
    r       = requests.get(url, headers=headers)
    if r.status_code == 200:
        d        = r.json()
        gen      = d.get('general', {})
        pulses   = d.get('pulse_info', {})
        return {
            'ip':          ip,
            'reputation':  gen.get('reputation', 'N/A'),
            'pulse_count': pulses.get('count', 0),
            'error':       ''
        }
    else:
        return {'ip': ip, 'reputation': 'N/A', 'pulse_count': 0,
                'error': r.text or f"HTTP {r.status_code}"}

# ── Kibana Login + Query ───────────────────────────────────────────────────────
session = requests.Session()
session.headers.update({'kbn-xsrf': 'true'})
session.auth = (KIBANA_USER, KIBANA_PASS)

def fetch_kibana_hits(size: int = PER_RUN_LIMIT) -> list:
    print(f"⏳ Fetching last 1 hr of alerts from Kibana (size={size})…")
    es_query = {
        "query": {
            "bool": {
                "filter": [
                    { "terms": { "client.id": CLIENT_IDS } },
                    {
                        "range": {
                            "@timestamp": {
                                "gte": "now-1h",
                                "format": "strict_date_optional_time"
                            }
                        }
                    }
                ]
            }
        },
        "size": size
    }
    proxy_path = f"/api/console/proxy?path={INDEX_PATTERN}/_search&method=POST"
    resp = session.post(KIBANA_URL.rstrip('/') + proxy_path, json=es_query)
    resp.raise_for_status()
    data    = resp.json()
    es_resp = data.get("responses", [data])[0]
    hits    = es_resp.get("hits", {}).get("hits", [])
    print(f"→ {len(hits)} total hits returned by Kibana")
    return hits

def extract_ips(hits: list) -> list:
    ips = set()
    for h in hits:
        src = h['_source'].get('source',{}).get('ip')
        dst = h['_source'].get('destination',{}).get('ip')
        for ip in (src, dst):
            if ip and is_public_ip(ip):
                ips.add(ip)
    ips = sorted(ips)[:PER_RUN_LIMIT]
    print(f"→ {len(ips)} unique public IPs to check (capped at {PER_RUN_LIMIT})")
    return ips

def write_to_excel(results: list, filename: str = 'ip_reputation_AV_OTX.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IP Reputation"
    ws.append(["IP", "Reputation", "Pulse Count", "Error"])
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for r in results:
        ws.append([r['ip'], r['reputation'], r['pulse_count'], r['error']])
        if r['pulse_count'] > 0 or r['error']:
            for c in ws[ws.max_row]:
                c.fill = red_fill
                c.font = Font(bold=True)

    wb.save(filename)
    print(f"→ Saved results to {filename}")

def main():
    print(f"▶️  CLIENT_IDS = {CLIENT_IDS}")
    hits = fetch_kibana_hits()
    ips  = extract_ips(hits)

    results = []
    for idx, ip in enumerate(ips, start=1):
        print(f"[{idx}/{len(ips)}] Checking {ip}…", end=' ')
        r = check_ip_otx(ip)
        print(f"pulses={r['pulse_count']}, rep={r['reputation']}, err={bool(r['error'])}")
        results.append(r)

        # simple rate-limit: pause every 50 requests to be nice
        if idx % 50 == 0:
            print("⏸️  Pausing 10 s to avoid hammering OTX…")
            time.sleep(10)

    write_to_excel(results)

if __name__ == "__main__":
    main()
